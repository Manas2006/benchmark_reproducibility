from fastapi import FastAPI, BackgroundTasks, WebSocket, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
import json
import asyncio
import os
import time
from pathlib import Path
import pandas as pd

from .schemas import EvalRequest, JobStatus, PathConfig, PathConfigResponse, CoTAnalysisResponse, CoTAnalysisResponseV2, OpenAITestRequest, OpenAITestResponse
from .runner import launch_job, job_db, get_job_status, cancel_job, delete_job, get_job_raw_data
from .schemas import EvalRequest, JobStatus, PathConfig, PathConfigResponse, CoTAnalysisResponse, TruncationAnalysisRequest, TruncationAnalysisResponse
from .runner import launch_job, job_db, get_job_status, cancel_job, delete_job, get_job_raw_data, save_job_db
from .path_manager import path_manager
from .cot_analyzer import CoTAnalyzer
import subprocess
import shlex
import requests

def _build_truncation_response(job_id: str, request: TruncationAnalysisRequest, output_dir: Path, computation_time: float) -> TruncationAnalysisResponse:
    """Helper function to build truncation analysis response"""
    model_stub = os.path.basename(request.model_name_or_path.rstrip('/'))
    raw_curves_path = output_dir / f"{request.dataset_name}_truncation_curves_{model_stub}.json"
    correct_plot_path = output_dir / f"{request.dataset_name}_correct_{model_stub}.png"
    incorrect_plot_path = output_dir / f"{request.dataset_name}_incorrect_{model_stub}.png"
    
    return TruncationAnalysisResponse(
        job_id=job_id,
        status="completed",
        message="Truncation analysis completed successfully",
        raw_curves_path=str(raw_curves_path) if raw_curves_path.exists() else None,
        correct_plot_path=str(correct_plot_path) if correct_plot_path.exists() else None,
        incorrect_plot_path=str(incorrect_plot_path) if incorrect_plot_path.exists() else None,
        computation_time=computation_time
    )

app = FastAPI(title="Qwen Math Evaluation API", version="1.0.0")

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Configure appropriately for production
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Mount static files for frontend
app.mount("/static", StaticFiles(directory="../frontend"), name="static")

@app.get("/")
async def root():
    """Serve the main frontend page"""
    return FileResponse("../frontend/index.html")

@app.get("/debug/html")
async def debug_html():
    """Debug endpoint to check HTML content"""
    try:
        with open("../frontend/index.html", "r") as f:
            content = f.read()
        return {"content_length": len(content), "has_cot_analysis": "cot-analysis" in content}
    except Exception as e:
        return {"error": str(e)}

@app.get("/health")
async def health_check():
    return {"status": "healthy"}

# Path configuration endpoints
@app.get("/config/paths")
async def get_path_config():
    """Get current path configuration"""
    current_config = path_manager.get_config()
    default_config = path_manager._get_default_config()
    return PathConfigResponse(
        current_config=current_config,
        default_config=default_config
    )

@app.post("/config/paths")
async def update_path_config(config: PathConfig):
    """Update path configuration"""
    path_manager.update_config(config)
    return {"message": "Path configuration updated successfully", "config": config}

@app.post("/config/paths/reset")
async def reset_path_config():
    """Reset path configuration to defaults"""
    path_manager.reset_to_default()
    return {"message": "Path configuration reset to defaults", "config": path_manager.get_config()}

@app.get("/config/paths/validate")
async def validate_paths():
    """Validate current path configuration"""
    validation = path_manager.validate_paths()
    return validation

@app.post("/jobs")
async def create_job(req: EvalRequest):
    """Create a new evaluation job"""
    jid = launch_job(req)
    # Return only serializable data
    job_info = job_db[jid].copy()
    job_info.pop("proc", None)
    job_info.pop("cli", None)
    # Add slurm_jid if present
    slurm_jid = job_info.get("slurm_jid")
    return {"job_id": jid, "slurm_jid": slurm_jid, **job_info}

@app.post("/jobs/{jid}/cancel")
async def cancel_job_endpoint(jid: str):
    """Cancel a running job (local or SLURM)"""
    success = cancel_job(jid)
    return {"job_id": jid, "cancelled": success}

@app.delete("/jobs/{jid}")
async def delete_job_endpoint(jid: str):
    """Delete a job (and cancel if running)"""
    success = delete_job(jid)
    return {"job_id": jid, "deleted": success}

@app.get("/jobs/{jid}")
async def job_status(jid: str):
    """Get the status of a job"""
    status_info = get_job_status(jid)
    if isinstance(status_info, dict):
        status_info = status_info.copy()
        status_info.pop("proc", None)
        status_info.pop("cli", None)
    slurm_jid = status_info.get("slurm_jid")
    result_file = status_info.get("result_file")
    prob_file = status_info.get("prob_file")
    return {"job_id": jid, "slurm_jid": slurm_jid, **status_info, "result_file": result_file, "prob_file": prob_file}

@app.get("/jobs/{jid}/prob-file")
async def get_prob_file(jid: str):
    if jid not in job_db:
        raise HTTPException(status_code=404, detail="Job not found")
    job_info = job_db[jid]
    prob_file = job_info.get("prob_file")
    if not prob_file:
        raise HTTPException(status_code=404, detail="Probability file not available for this job")
    file_path = Path(prob_file)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Probability file not found on disk")
    # Security check
    config = path_manager.get_config()
    allowed_dirs = [Path(config.output_dir), Path(config.logs_dir), Path(config.evaluation_dir)]
    is_allowed = False
    for allowed_dir in allowed_dirs:
        try:
            file_path.relative_to(allowed_dir)
            is_allowed = True
            break
        except ValueError:
            continue
    if not is_allowed:
        raise HTTPException(status_code=403, detail="Access denied to this file path")
    return FileResponse(path=str(file_path), filename=file_path.name, media_type='application/json')

@app.get("/jobs/{jid}/prob-plot")
async def generate_prob_plot(jid: str, plot_type: str, sample_id: int | None = None, math_level: str | None = None):
    valid_plot_types = ("aggregate", "single", "path_aggregate", "path_single", "correct_aggregate", "incorrect_aggregate", 
                       "level_single", "level_aggregate", "starting_tokens_by_level", "ending_tokens_by_level", "correct_vs_incorrect")
    if plot_type not in valid_plot_types:
        raise HTTPException(status_code=400, detail=f"plot_type must be one of: {', '.join(valid_plot_types)}")
    if jid not in job_db:
        raise HTTPException(status_code=404, detail="Job not found")
    job_info = job_db[jid]
    prob_file = job_info.get("prob_file")
    if not prob_file:
        raise HTTPException(status_code=404, detail="Probability file not available for this job")
    prob_path = Path(prob_file)
    if not prob_path.exists():
        raise HTTPException(status_code=404, detail="Probability file not found on disk")

    # Build plotting command
    config = path_manager.get_config()
    python_bin = config.python_path
    eval_dir = Path(config.evaluation_dir)
    plot_script = eval_dir / "prob_plot.py"
    if not plot_script.exists():
        raise HTTPException(status_code=500, detail=f"Plot script not found: {plot_script}")

    # Names
    dataset_name = job_info.get("request", {}).get("dataset", "dataset")
    model_name = job_info.get("request", {}).get("model", "model")
    prompt_type = job_info.get("request", {}).get("prompt_type", "") or "custom"
    method_name = f"{Path(model_name).name}-{prompt_type}"

    output_dir = prob_path.parent / "prob_plots"
    output_dir.mkdir(parents=True, exist_ok=True)

    # Expected output path
    if plot_type == "aggregate":
        expected_png = output_dir / f"{dataset_name}_aggregate_{method_name}.png"
    elif plot_type == "correct_aggregate":
        expected_png = output_dir / f"{dataset_name}_correct_aggregate_{method_name}.png"
    elif plot_type == "incorrect_aggregate":
        expected_png = output_dir / f"{dataset_name}_incorrect_aggregate_{method_name}.png"
    elif plot_type == "correct_vs_incorrect":
        expected_png = output_dir / f"{dataset_name}_correct_vs_incorrect_{method_name}.png"
    elif plot_type == "path_aggregate":
        expected_png = output_dir / f"{dataset_name}_path_aggregate_{method_name}.png"
    elif plot_type == "path_single":
        if sample_id is None:
            raise HTTPException(status_code=400, detail="sample_id is required for path_single plot")
        expected_png = output_dir / f"{dataset_name}_path_single_{method_name}_id_{sample_id}.png"
    elif plot_type == "level_single":
        if math_level is None:
            raise HTTPException(status_code=400, detail="math_level is required for level_single plot")
        expected_png = output_dir / f"{dataset_name}_level_{math_level}_{method_name}.png"
    elif plot_type == "level_aggregate":
        # For level_aggregate, we'll return multiple images as a zip file
        expected_png = output_dir / f"{dataset_name}_level_aggregate_{method_name}.zip"
    elif plot_type == "starting_tokens_by_level":
        expected_png = output_dir / f"{dataset_name}_starting_tokens_by_level_{method_name}.png"
    elif plot_type == "ending_tokens_by_level":
        expected_png = output_dir / f"{dataset_name}_ending_tokens_by_level_{method_name}.png"
    else:  # single
        if sample_id is None:
            raise HTTPException(status_code=400, detail="sample_id is required for single plot")
        expected_png = output_dir / f"{dataset_name}_single_{method_name}_id_{sample_id}.png"

    cmd = [
        python_bin,
        str(plot_script),
        str(prob_path),
        "--dataset_name", str(dataset_name),
        "--method_name", str(method_name),
        "--output_dir", str(output_dir),
        "--plot_type", plot_type,
    ]
    if plot_type in ("single", "path_single"):
        cmd.extend(["--sample_id", str(sample_id)])
    elif plot_type == "level_single":
        cmd.extend(["--math_level", str(math_level)])

    try:
        result = subprocess.run(cmd, capture_output=True, text=True, cwd=eval_dir)
        if result.returncode != 0:
            raise HTTPException(status_code=500, detail=f"Plotting failed: {result.stderr}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error running plot script: {str(e)}")

    if not expected_png.exists():
        # Fallback: try to find any recent file in output_dir
        if plot_type == "level_aggregate":
            # Look for zip files for level_aggregate
            files = sorted(output_dir.glob("*.zip"), key=lambda p: p.stat().st_mtime, reverse=True)
            if not files:
                raise HTTPException(status_code=500, detail="Level aggregate plot archive not generated")
            expected_png = files[0]
        else:
            # Look for PNG files for other plot types
            files = sorted(output_dir.glob("*.png"), key=lambda p: p.stat().st_mtime, reverse=True)
            if not files:
                raise HTTPException(status_code=500, detail="Plot image not generated")
            expected_png = files[0]

    # Return appropriate media type based on file extension
    if expected_png.suffix.lower() == '.zip':
        return FileResponse(path=str(expected_png), filename=expected_png.name, media_type='application/zip')
    else:
        return FileResponse(path=str(expected_png), filename=expected_png.name, media_type='image/png')

@app.websocket("/stream/{jid}")
async def stream(jid: str, ws: WebSocket):
    await ws.accept()
    info = job_db.get(jid)
    if not info:
        try:
            await ws.send_text(json.dumps({"error": f"Job {jid} not found (UUID or SLURM job number)"}))
        except Exception:
            pass
        try:
            await ws.close()
        except Exception:
            pass
        return

    import os
    import asyncio
    out_path = info.get("out_file")
    err_path = info.get("err_file")
    # For SLURM jobs, if the files do not exist, try to parse the SBATCH script for the real paths
    if info.get("backend") == "slurm":
        sbatch_path = info.get("sbatch_path")
        slurm_jid = info.get("slurm_jid")
        if sbatch_path and (not out_path or not err_path):
            try:
                with open(sbatch_path, "r") as f:
                    lines = f.readlines()
                for line in lines:
                    if line.startswith("#SBATCH -o"):
                        out_path = line.split(None, 2)[-1].strip()
                    if line.startswith("#SBATCH -e"):
                        err_path = line.split(None, 2)[-1].strip()
                
                # Replace %j placeholder with actual SLURM job ID
                if slurm_jid and out_path and "%j" in out_path:
                    out_path = out_path.replace("%j", str(slurm_jid))
                if slurm_jid and err_path and "%j" in err_path:
                    err_path = err_path.replace("%j", str(slurm_jid))
            except Exception:
                pass
    # Wait for the files to appear (up to 60 seconds)
    wait_time = 0
    while (not out_path or not os.path.exists(out_path) or not err_path or not os.path.exists(err_path)) and wait_time < 60:
        try:
            await ws.send_text(json.dumps({"waiting": f"Waiting for job output files to appear... ({wait_time}s)"}))
        except Exception:
            # WebSocket connection lost during waiting
            return
        await asyncio.sleep(2)
        wait_time += 2
    if not out_path or not os.path.exists(out_path) or not err_path or not os.path.exists(err_path):
        try:
            await ws.send_text(json.dumps({"error": f"Output or error file not found for this job after waiting. (out: {out_path}, err: {err_path})"}))
        except Exception:
            pass
        try:
            await ws.close()
        except Exception:
            pass
        return
    try:
        with open(out_path, "r") as outf, open(err_path, "r") as errf:
            outf.seek(0, os.SEEK_END)
            errf.seek(0, os.SEEK_END)
            while True:
                out_line = outf.readline()
                err_line = errf.readline()
                sent = False
                if out_line:
                    try:
                        line = out_line.rstrip()
                        # Check for structured monitoring information
                        if line.startswith("MONITOR_PROMPT:"):
                            try:
                                monitor_data = json.loads(line[15:])  # Remove "MONITOR_PROMPT: "
                                await ws.send_text(json.dumps({"monitor_prompt": monitor_data}))
                            except:
                                await ws.send_text(json.dumps({"out": line}))
                        elif line.startswith("MONITOR_EPOCH:"):
                            try:
                                monitor_data = json.loads(line[14:])  # Remove "MONITOR_EPOCH: "
                                await ws.send_text(json.dumps({"monitor_epoch": monitor_data}))
                            except:
                                await ws.send_text(json.dumps({"out": line}))
                        elif line.startswith("MONITOR_RESPONSE:"):
                            try:
                                monitor_data = json.loads(line[17:])  # Remove "MONITOR_RESPONSE: "
                                await ws.send_text(json.dumps({"monitor_response": monitor_data}))
                            except:
                                await ws.send_text(json.dumps({"out": line}))
                        else:
                            await ws.send_text(json.dumps({"out": line}))
                        sent = True
                    except Exception:
                        # WebSocket connection lost
                        break
                if err_line:
                    try:
                        await ws.send_text(json.dumps({"err": err_line.rstrip()}))
                        sent = True
                    except Exception:
                        # WebSocket connection lost
                        break
                # Check if job is still running
                status = info.get("status")
                if status not in ["RUNNING", "QUEUED"]:
                    # Drain any remaining lines
                    try:
                        for line in outf:
                            await ws.send_text(json.dumps({"out": line.rstrip()}))
                        for line in errf:
                            await ws.send_text(json.dumps({"err": line.rstrip()}))
                    except Exception:
                        # WebSocket connection lost during drain
                        pass
                    break
                if not sent:
                    await asyncio.sleep(1)
    except Exception as e:
        try:
            await ws.send_text(json.dumps({"error": str(e)}))
        except Exception:
            # WebSocket connection already closed
            pass
    finally:
        try:
            await ws.close()
        except Exception:
            # WebSocket already closed
            pass
    return

@app.get("/file")
async def serve_file(path: str):
    """Serve a file from the filesystem"""
    try:
        # Decode the URL-encoded path
        import urllib.parse
        decoded_path = urllib.parse.unquote(path)
        
        # Convert to Path object for better handling
        file_path = Path(decoded_path)
        
        # Security check: ensure the path is within allowed directories
        config = path_manager.get_config()
        allowed_dirs = [
            Path(config.output_dir),
            Path(config.logs_dir),
            Path(config.evaluation_dir),
        ]
        
        is_allowed = False
        for allowed_dir in allowed_dirs:
            try:
                file_path.relative_to(allowed_dir)
                is_allowed = True
                break
            except ValueError:
                continue
        
        if not is_allowed:
            raise HTTPException(status_code=403, detail="Access denied to this file path")
        
        # Check if file exists
        if not file_path.exists():
            # Check if this might be a job result file that hasn't been generated yet
            if "test_" in file_path.name and file_path.suffix in [".jsonl", ".json"]:
                raise HTTPException(
                    status_code=404, 
                    detail=f"Result file not found: {file_path.name}. The job may still be running or may have failed."
                )
            else:
                raise HTTPException(status_code=404, detail=f"File not found: {file_path}")
        
        # Check if it's actually a file
        if not file_path.is_file():
            raise HTTPException(status_code=400, detail=f"Path is not a file: {file_path}")
        
        # Return the file
        return FileResponse(
            path=str(file_path),
            filename=file_path.name,
            media_type='application/octet-stream'
        )
        
    except HTTPException:
        # Re-raise HTTP exceptions as-is
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error serving file: {str(e)}")

@app.get("/metrics/{job_id}")
async def get_metrics_file(job_id: str):
    """Find and serve metrics file for a specific job"""
    try:
        if job_id not in job_db:
            raise HTTPException(status_code=404, detail="Job not found")
        
        job_info = job_db[job_id]
        result_file = job_info.get("result_file")
        
        if not result_file:
            raise HTTPException(status_code=404, detail="No result file found for this job")
        
        # Convert result file path to Path object
        result_path = Path(result_file)
        
        # Security check
        config = path_manager.get_config()
        allowed_dirs = [Path(config.output_dir), Path(config.logs_dir), Path(config.evaluation_dir)]
        
        is_allowed = False
        for allowed_dir in allowed_dirs:
            try:
                result_path.relative_to(allowed_dir)
                is_allowed = True
                break
            except ValueError:
                continue
        
        if not is_allowed:
            raise HTTPException(status_code=403, detail="Access denied to this file path")
        
        # Try different patterns for metrics file
        base_name = result_path.stem  # Remove extension
        prompt_type = job_info.get("request", {}).get("prompt_type", "cot")
        
        possible_metrics_files = [
            result_path.parent / f"{base_name}_{prompt_type}_metrics.json",
            result_path.parent / f"{base_name}_metrics.json",
            result_path.parent / f"{result_path.stem}_{prompt_type}_metrics.json",
            result_path.parent / f"{result_path.stem}_metrics.json"
        ]
        
        # Find the first existing metrics file
        metrics_file = None
        for possible_file in possible_metrics_files:
            if possible_file.exists():
                metrics_file = possible_file
                break
        
        if not metrics_file:
            raise HTTPException(
                status_code=404, 
                detail=f"Metrics file not found for job {job_id}. The job may still be running or may have failed."
            )
        
        # Return the metrics file
        return FileResponse(
            path=str(metrics_file),
            filename=metrics_file.name,
            media_type='application/json'
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error finding metrics file: {str(e)}")

@app.get("/jobs")
async def list_jobs():
    jobs = []
    try:
        for jid, info in job_db.items():
            # Get real-time status for each job
            job_info = get_job_status(jid)
            if isinstance(job_info, dict):
                job_info = job_info.copy()
                job_info.pop("proc", None)
                job_info.pop("cli", None)
            slurm_jid = job_info.get("slurm_jid")
            result_file = job_info.get("result_file")
            prob_file = job_info.get("prob_file")
            jobs.append({"job_id": jid, "slurm_jid": slurm_jid, **job_info, "result_file": result_file, "prob_file": prob_file})
        return {"jobs": jobs}
    except Exception as e:
        print(f"Error in /jobs: {e}")
        return {"jobs": [], "error": str(e)}

def export_results_to_excel(job_id: str):
    """Export evaluation results to Excel file"""
    try:
        if job_id not in job_db:
            raise HTTPException(status_code=404, detail="Job not found")
        
        job_info = job_db[job_id]
        result_file = job_info.get("result_file")
        
        if not result_file:
            raise HTTPException(status_code=404, detail="No result file found for this job")
        
        # Convert result file path to Path object
        result_path = Path(result_file)
        
        # Security check
        config = path_manager.get_config()
        allowed_dirs = [Path(config.output_dir), Path(config.logs_dir), Path(config.evaluation_dir)]
        
        is_allowed = False
        for allowed_dir in allowed_dirs:
            try:
                result_path.relative_to(allowed_dir)
                is_allowed = True
                break
            except ValueError:
                continue
        
        if not is_allowed:
            raise HTTPException(status_code=403, detail="Access denied to this file path")
        
        # Load the JSON results (JSON Lines format)
        if not result_path.exists():
            raise HTTPException(status_code=404, detail="Result file not found")
        
        data = []
        with open(result_path, 'r') as f:
            for line in f:
                line = line.strip()
                if line:  # Skip empty lines
                    data.append(json.loads(line))
        
        if len(data) == 0:
            raise HTTPException(status_code=400, detail="Invalid result file format or empty file")
        
        # Initialize CoT analyzer for metrics calculation
        config = path_manager.get_config()
        analyzer = CoTAnalyzer(openai_api_key=config.openai_api_key)
        
        # Extract model name and configuration from result file path
        model_name = "Unknown"
        model_config = "Unknown"
        if result_file:
            # Extract model name from path like: .../Qwen2.5-Math-1.5B/gsm8k/test_auto-cot_-1_seed0_t0.0_s0_e-1_...
            path_parts = Path(result_file).parts
            for part in path_parts:
                if any(model in part.lower() for model in ['qwen', 'gpt', 'claude', 'llama', 'mistral', 'gemini']):
                    model_name = part
                    break
            
            # Extract configuration from filename
            filename = Path(result_file).stem
            if 't0.0' in filename:
                model_config = "Temperature: 0.0"
            elif 't0.1' in filename:
                model_config = "Temperature: 0.1"
            elif 't0.7' in filename:
                model_config = "Temperature: 0.7"
            elif 't1.0' in filename:
                model_config = "Temperature: 1.0"
            
            # Extract other config details
            config_parts = []
            if 'seed0' in filename:
                config_parts.append("Seed: 0")
            if 'auto-cot' in filename:
                config_parts.append("Auto-CoT: Enabled")
            if config_parts:
                model_config += f", {', '.join(config_parts)}"

        # Build DataFrame with all available columns + CoT metrics
        df_data = []
        all_cot_metrics = []  # Store metrics for summary calculation
        
        for rec in data:
            # Handle list fields by converting to string
            code_list = rec.get('code', [])
            pred_list = rec.get('pred', [])
            score_list = rec.get('score', [])
            
            # Calculate CoT metrics for this answer
            # Use the model's actual output (code field) + predicted answer, not the formatted answer field
            model_reasoning = rec.get('code', [''])
            model_reasoning_text = model_reasoning[0] if model_reasoning else ""
            predicted_answer = rec.get('pred', [''])
            predicted_answer_text = predicted_answer[0] if predicted_answer else ""
            
            # Construct the full model output for CoT analysis
            full_model_output = f"{model_reasoning_text}\n#### {predicted_answer_text}"
            
            gt = rec.get('gt', '')
            cot_metrics = analyzer.analyze_answer(full_model_output, gt)
            all_cot_metrics.append(cot_metrics)
            
            # Build row data with original columns + CoT metrics
            # Construct the prompt if it's missing but we have the question
            prompt_text = rec.get('prompt', '')
            if not prompt_text and rec.get('question'):
                # Reconstruct the basic CoT prompt format
                prompt_text = f"Question: {rec.get('question', '')}\nAnswer:"
            
            # Generate GPT prompt for this sample (if we have the CoT analysis)
            gpt_prompt = ""
            try:
                # Try to get CoT analysis for this sample to extract the GPT prompt
                cot_response = requests.get(f'http://localhost:8001/jobs/{job_id}/cot-analysis')
                if cot_response.status_code == 200:
                    cot_data = cot_response.json()
                    sample_idx = rec.get('idx', 0)
                    if sample_idx < len(cot_data.get('per_sample', [])):
                        sample_data = cot_data['per_sample'][sample_idx]
                        # Extract the GPT prompt from the judge_raw or build it
                        if 'judge_raw' in sample_data:
                            # If we have judge data, we can reconstruct the prompt
                            problem = sample_data.get('problem', '')
                            model_output = sample_data.get('model_output', '')
                            gold = sample_data.get('gold', '')
                            flags = sample_data.get('flags', [])
                            evidence = sample_data.get('evidence', {})
                            
                            # Build a simplified version of the GPT prompt
                            gpt_prompt = f"""Problem: {problem}

Model Reasoning: {model_output}

Gold Answer: {gold}

Flags: {flags}

Evidence: {evidence}"""
            except:
                gpt_prompt = "GPT prompt not available"

            row_data = {
                'Index': rec.get('idx', ''),
                'Model_Name': model_name,
                'Model_Configuration': model_config,
                'GPT_Prompt': gpt_prompt,
                'Prompt': prompt_text,
                'Model Code Output': '\n'.join(str(item) for item in code_list if item is not None),
                'Prediction': '\n'.join(str(item) for item in pred_list if item is not None),
                'Ground Truth': rec.get('gt', ''),
                'Score': '\n'.join(str(item) for item in score_list if item is not None),
                
                # CoT Analysis Metrics - CQS Components
                'CQS_Final_Answer_Correct': round(cot_metrics.final_answer_correctness, 3),
                'CQS_Arithmetic_Accuracy': round(cot_metrics.arithmetic_accuracy, 3),
                'CQS_Logical_Structure': round(cot_metrics.logical_structure_score, 3),
                'CQS_Consistency_Complete': round(cot_metrics.consistency_completeness, 3),
                'CQS_Formatting_Notation': round(cot_metrics.formatting_notation, 3),
                'CQS_Overall_Score': round(cot_metrics.cqs_score, 3),
                
                # Basic CoT Metrics
                'CoT_Reasoning_Steps': cot_metrics.reasoning_steps,
                'CoT_Total_Characters': cot_metrics.total_chars,
                'CoT_Avg_Words_Per_Step': round(cot_metrics.avg_words_per_step, 2),
                'CoT_Arithmetic_Expressions': cot_metrics.arithmetic_expressions,
                
                # Legacy Boolean Metrics
                'CoT_Has_Clear_Structure': cot_metrics.has_clear_structure,
                'CoT_Has_Final_Answer': cot_metrics.has_final_answer,
                'CoT_Uses_Calculations': cot_metrics.uses_intermediate_calculations,
                'CoT_Shows_Work': cot_metrics.shows_work_explicitly,
                'CoT_Logical_Sequence': cot_metrics.follows_logical_sequence,
                'CoT_Error_Patterns': ', '.join(cot_metrics.error_patterns) if cot_metrics.error_patterns else '',
                'CoT_Confidence_Score': round(cot_metrics.confidence_score, 3)
            }
            
            df_data.append(row_data)
        
        df = pd.DataFrame(df_data)
        
        # Calculate summary statistics for CoT metrics
        if all_cot_metrics:
            total_samples = len(all_cot_metrics)
            # Basic metrics
            avg_reasoning_steps = sum(m.reasoning_steps for m in all_cot_metrics) / total_samples
            avg_total_chars = sum(m.total_chars for m in all_cot_metrics) / total_samples
            avg_words_per_step = sum(m.avg_words_per_step for m in all_cot_metrics) / total_samples
            
            # CQS Component Averages
            avg_final_answer_correct = sum(m.final_answer_correctness for m in all_cot_metrics) / total_samples
            avg_arithmetic_acc = sum(m.arithmetic_accuracy for m in all_cot_metrics) / total_samples
            avg_logical_structure = sum(m.logical_structure_score for m in all_cot_metrics) / total_samples
            avg_consistency_complete = sum(m.consistency_completeness for m in all_cot_metrics) / total_samples
            avg_formatting_notation = sum(m.formatting_notation for m in all_cot_metrics) / total_samples
            avg_cqs_score = sum(m.cqs_score for m in all_cot_metrics) / total_samples
            
            # Legacy metrics
            avg_confidence = sum(m.confidence_score for m in all_cot_metrics) / total_samples
            
            # Pattern counts
            clear_structure_count = sum(1 for m in all_cot_metrics if m.has_clear_structure)
            final_answer_count = sum(1 for m in all_cot_metrics if m.has_final_answer)
            uses_calc_count = sum(1 for m in all_cot_metrics if m.uses_intermediate_calculations)
            shows_work_count = sum(1 for m in all_cot_metrics if m.shows_work_explicitly)
            logical_seq_count = sum(1 for m in all_cot_metrics if m.follows_logical_sequence)
            
            # Error pattern frequency
            error_patterns = {}
            for m in all_cot_metrics:
                for error in m.error_patterns:
                    error_patterns[error] = error_patterns.get(error, 0) + 1
        
        # Create Excel file path
        excel_path = result_path.parent / f"{result_path.stem}_results.xlsx"
        
        # Write to Excel with multiple sheets
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Sheet 1: Detailed results with per-row CoT metrics
            df.to_excel(writer, index=False, sheet_name='Detailed Results')
            
            # Sheet 2: Model Information
            model_info_data = [
                ['Model Information', ''],
                ['', ''],
                ['Model Name', model_name],
                ['Model Configuration', model_config],
                ['Total Samples', len(data)],
                ['Result File', result_file],
                ['Export Date', pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')],
                ['', ''],
                ['Evaluation Settings', ''],
                ['CoT Analysis', 'Enabled'],
                ['GPT Judge Mode', 'ALWAYS'],
                ['Flag Detection', 'Automated'],
                ['Scoring Method', 'Hybrid (Rule-based + LLM)']
            ]
            
            model_info_df = pd.DataFrame(model_info_data, columns=['Setting', 'Value'])
            model_info_df.to_excel(writer, index=False, sheet_name='Model Info')
            
            # Sheet 3: CoT Analysis Summary
            if all_cot_metrics:
                summary_data = [
                    ['CoT Quality Score (CQS) Analysis', ''],
                    ['', ''],
                    ['=== OVERALL METRICS ===', ''],
                    ['Total Samples', total_samples],
                    ['Average CQS Score', round(avg_cqs_score, 3)],
                    ['', ''],
                    ['=== CQS COMPONENT SCORES ===', ''],
                    ['Final Answer Correctness (30%)', round(avg_final_answer_correct, 3)],
                    ['Arithmetic Accuracy (25%)', round(avg_arithmetic_acc, 3)],
                    ['Logical Structure (20%)', round(avg_logical_structure, 3)],
                    ['Consistency & Completeness (15%)', round(avg_consistency_complete, 3)],
                    ['Formatting & Notation (10%)', round(avg_formatting_notation, 3)],
                    ['', ''],
                    ['=== BASIC METRICS ===', ''],
                    ['Average Reasoning Steps', round(avg_reasoning_steps, 2)],
                    ['Average Character Count', round(avg_total_chars, 2)],
                    ['Average Words per Step', round(avg_words_per_step, 2)],
                    ['Average Confidence Score', round(avg_confidence, 3)],
                    ['', ''],
                    ['=== PATTERN ANALYSIS ===', ''],
                    ['Samples with Clear Structure', f"{clear_structure_count} ({clear_structure_count/total_samples*100:.1f}%)"],
                    ['Samples with Final Answer', f"{final_answer_count} ({final_answer_count/total_samples*100:.1f}%)"],
                    ['Samples Using Calculations', f"{uses_calc_count} ({uses_calc_count/total_samples*100:.1f}%)"],
                    ['Samples Showing Work', f"{shows_work_count} ({shows_work_count/total_samples*100:.1f}%)"],
                    ['Samples with Logical Sequence', f"{logical_seq_count} ({logical_seq_count/total_samples*100:.1f}%)"],
                    ['', ''],
                    ['=== ERROR PATTERNS ===', '']
                ]
                
                # Add error patterns to summary
                for error, count in error_patterns.items():
                    summary_data.append([f"  {error}", f"{count} ({count/total_samples*100:.1f}%)"])
                
                summary_df = pd.DataFrame(summary_data, columns=['Metric', 'Value'])
                summary_df.to_excel(writer, index=False, sheet_name='CoT Summary')
            
            # Get the workbook and worksheets for formatting
            workbook = writer.book
            results_worksheet = writer.sheets['Detailed Results']
            
            # Create fill styles for conditional formatting
            from openpyxl.styles import PatternFill, Font
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            blue_fill = PatternFill(start_color='D4E6F1', end_color='D4E6F1', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            
            # Find important columns for conditional formatting
            score_col = None
            clarity_col = None
            confidence_col = None
            
            for col_idx, col_name in enumerate(df.columns, 1):
                if col_name == 'Score':
                    score_col = col_idx
                elif col_name == 'CoT_Clarity_Score':
                    clarity_col = col_idx
                elif col_name == 'CoT_Confidence_Score':
                    confidence_col = col_idx
            
            # Apply conditional formatting to multiple columns
            for row_idx in range(2, len(df) + 2):  # Start from row 2 (after header)
                # Score column (green/red for true/false)
                if score_col:
                    cell = results_worksheet.cell(row=row_idx, column=score_col)
                    if cell.value and 'true' in str(cell.value).lower():
                        cell.fill = green_fill
                    elif cell.value and 'false' in str(cell.value).lower():
                        cell.fill = red_fill
                
                # Clarity score column (blue gradient based on score)
                if clarity_col:
                    cell = results_worksheet.cell(row=row_idx, column=clarity_col)
                    if cell.value and isinstance(cell.value, (int, float)):
                        if cell.value >= 0.8:
                            cell.fill = blue_fill
                        elif cell.value <= 0.5:
                            cell.fill = yellow_fill
                
                # Confidence score column (similar to clarity)
                if confidence_col:
                    cell = results_worksheet.cell(row=row_idx, column=confidence_col)
                    if cell.value and isinstance(cell.value, (int, float)):
                        if cell.value >= 0.8:
                            cell.fill = blue_fill
                        elif cell.value <= 0.5:
                            cell.fill = yellow_fill
            
            # Format the summary sheet if it exists
            if all_cot_metrics and 'CoT Summary' in writer.sheets:
                summary_worksheet = writer.sheets['CoT Summary']
                
                # Make the title bold
                title_cell = summary_worksheet.cell(row=1, column=1)
                title_cell.font = Font(bold=True, size=14)
                
                # Make section headers bold
                for row_idx in range(1, len(summary_data) + 1):
                    cell = summary_worksheet.cell(row=row_idx, column=1)
                    if cell.value in ['Pattern Analysis', 'Error Patterns']:
                        cell.font = Font(bold=True)
        
        return str(excel_path)
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting to Excel: {str(e)}")

@app.get('/jobs/{job_id}/export')
async def export_job_excel(job_id: str):
    """Export job results to Excel file"""
    try:
        excel_path = export_results_to_excel(job_id)
        excel_file = Path(excel_path)
        
        return FileResponse(
            path=str(excel_file),
            filename=excel_file.name,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error serving Excel file: {str(e)}")

@app.get('/jobs/{job_id}/raw-answers')
async def get_job_raw_answers(job_id: str):
    """Get raw answer data for CoT analysis"""
    try:
        raw_data = get_job_raw_data(job_id)
        return raw_data
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching raw data: {str(e)}")

@app.get("/jobs/{job_id}/data-health")
async def check_job_data_health(job_id: str):
    """Health check for job data availability"""
    try:
        raw_data = get_job_raw_data(job_id)
        data_list = raw_data.get('data', [])
        return {
            "job_id": job_id,
            "data_available": True,
            "sample_count": len(data_list),
            "has_answers": sum(1 for r in data_list if r.get('answer')),
            "avg_answer_length": sum(len(r.get('answer', '')) for r in data_list) / len(data_list) if data_list else 0,
            "metadata": raw_data.get('metadata', {})
        }
    except Exception as e:
        return {
            "job_id": job_id,
            "data_available": False,
            "error": str(e)
        }

@app.get("/jobs/{job_id}/cot-analysis", response_model=CoTAnalysisResponseV2)
async def get_cot_analysis(job_id: str):
    """Get Chain-of-Thought analysis for a job using new four-pillar evaluation"""
    start_time = time.time()
    
    try:
        # Get raw data
        raw_data = get_job_raw_data(job_id)
        data_list = raw_data.get('data', [])
        
        if not data_list:
            raise HTTPException(status_code=404, detail="No data found for analysis")
        
        # Initialize new Pillars v2 evaluator
        from .cot_eval_v2.evaluator import PillarsEvaluator
        from .cot_eval_v2.judge import Judge
        
        # Get configuration
        config = path_manager.get_config()
        
        # Create judge if OpenAI key available
        judge = None
        if config.openai_api_key:
            # Set the API key as environment variable for OpenAI client
            import os
            os.environ['OPENAI_API_KEY'] = config.openai_api_key
            judge = Judge(mode="ALWAYS")
        
        # Initialize evaluator
        evaluator = PillarsEvaluator(judge=judge)
        
        # Process each sample
        results = []
        total_samples = len(data_list)
        
        print(f"🔍 Starting CoT analysis for {total_samples} samples...")
        
        for i, sample in enumerate(data_list):
            try:
                print(f"📊 Analyzing sample {i+1}/{total_samples}...")
                
                # Extract model reasoning and answer
                model_reasoning = sample.get('code', [''])
                model_reasoning_text = model_reasoning[0] if model_reasoning else ""
                predicted_answer = sample.get('pred', [''])
                predicted_answer_text = predicted_answer[0] if predicted_answer else ""
                
                # Construct full CoT text
                full_cot_text = f"{model_reasoning_text}\n#### {predicted_answer_text}"
                
                # Get existing correctness from job data
                existing_scores = sample.get('score', [])
                is_correct = existing_scores[0] if existing_scores else False
                
                print(f"   Question: {sample.get('question', '')[:100]}{'...' if len(sample.get('question', '')) > 100 else ''}")
                print(f"   Ground Truth: {sample.get('gt', '')}")
                print(f"   Predicted: {predicted_answer_text}")
                print(f"   Correct: {is_correct}")
                
                # Run analysis
                print(f"   🔍 Running Pillars v2 evaluation...")
                flags, evidence, rule_scores, judge_scores, fused_scores = evaluator.analyze(
                    problem=sample.get('question', ''),
                    cot_text=full_cot_text,
                    gold=sample.get('gt', '')
                )
                
                # Override the final_correct with the existing job data
                evidence['final_correct'] = is_correct
                
                # Recalculate rule scores with the correct final_correct value
                from .cot_eval_v2.scoring import rule_scores
                rule_scores = rule_scores(evidence)
                
                # Recalculate fused scores
                from .cot_eval_v2.scoring import fuse_with_judge
                fused_scores = fuse_with_judge(rule_scores, judge_scores, evidence)
                
                # Convert flags to summary format
                flags_summary = flags.summarize_for_prompt()
                
                # Convert flags to dictionary format for API response
                flags_dict = {}
                for pillar in ["faithfulness", "utility", "coherence", "factuality"]:
                    pillar_flags = flags.get_flags_by_pillar(pillar)
                    flags_dict[pillar] = [flag.to_dict() for flag in pillar_flags]
                
                # Log analysis results
                print(f"   ✅ Analysis complete!")
                print(f"   📊 Rule scores: {rule_scores}")
                print(f"   🤖 Judge scores: {judge_scores}")
                print(f"   🔗 Fused scores: {fused_scores}")
                total_flags = sum(len(flags_dict[pillar]) for pillar in flags_dict)
                print(f"   🚩 Flags found: {total_flags}")
                if total_flags > 0:
                    for pillar in ["faithfulness", "utility", "coherence", "factuality"]:
                        pillar_flags = flags_dict[pillar]
                        if pillar_flags:
                            print(f"      - {pillar}: {len(pillar_flags)} flags")
                            for flag in pillar_flags[:2]:  # Show first 2 flags per pillar
                                print(f"        * {flag['issue']} (step: {flag['step']})")
                print()
                
                # Build result for this sample
                sample_result = {
                    "sample_id": i,
                    "problem": sample.get('question', ''),
                    "cot_text": full_cot_text,
                    "gold_answer": sample.get('gt', ''),
                    "flags": flags_dict,
                    "evidence": evidence,
                    "rule_scores": rule_scores,
                    "judge_scores": judge_scores,
                    "fused_scores": fused_scores,
                    "final_answer": predicted_answer_text
                }
                
                results.append(sample_result)
                
            except Exception as e:
                print(f"Error processing sample {i}: {e}")
                # Add error result
                results.append({
                    "sample_id": i,
                    "problem": sample.get('question', ''),
                    "cot_text": "",
                    "gold_answer": sample.get('gt', ''),
                    "error": str(e),
                    "flags": {},
                    "evidence": {},
                    "rule_scores": {},
                    "judge_scores": {},
                    "fused_scores": {},
                    "final_answer": ""
                })
        
        # Calculate summary statistics
        valid_results = [r for r in results if "error" not in r]
        if valid_results:
            # Calculate average scores
            avg_scores = {}
            for pillar in ["faithfulness", "utility", "coherence", "factuality"]:
                scores = [r["fused_scores"].get(pillar, 0) for r in valid_results]
                avg_scores[pillar] = sum(scores) / len(scores) if scores else 0
            
            avg_scores["overall"] = sum(avg_scores.values()) / len(avg_scores)
            
            # Count flags
            flag_counts = {}
            for result in valid_results:
                for flag_type, flags in result["flags"].items():
                    if flag_type not in flag_counts:
                        flag_counts[flag_type] = 0
                    flag_counts[flag_type] += len(flags)
        else:
            avg_scores = {"faithfulness": 0, "utility": 0, "coherence": 0, "factuality": 0, "overall": 0}
            flag_counts = {}
        
        # Convert results to proper schema format
        per_sample = []
        for result in results:
            if "error" not in result:
                # Convert flags to PillarsFlag format
                flags_list = []
                for pillar, flags in result["flags"].items():
                    for flag in flags:
                        flags_list.append({
                            "pillar": pillar,
                            "step": flag.get("step", "unknown"),
                            "issue": flag.get("issue", "unknown"),
                            "details": flag.get("details", {})
                        })
                
                # Create PillarsScores
                scores = {
                    "faithfulness": result["fused_scores"].get("faithfulness", 0.0),
                    "utility": result["fused_scores"].get("utility", 0.0),
                    "coherence": result["fused_scores"].get("coherence", 0.0),
                    "factuality": result["fused_scores"].get("factuality", 0.0),
                    "overall": result["fused_scores"].get("overall", 0.0)
                }
                
                per_sample.append({
                    "scores": scores,
                    "flags": flags_list,
                    "evidence": result["evidence"],
                    "rules_raw": result["rule_scores"],
                    "judge_raw": result["judge_scores"],
                    "config_snapshot": {"judge_available": judge is not None},
                    "problem": result["problem"],
                    "model_output": result["cot_text"],
                    "gold": result["gold_answer"]
                })
        
        # Create summary
        summary = {
            "avg_faithfulness": avg_scores.get("faithfulness", 0.0),
            "avg_utility": avg_scores.get("utility", 0.0),
            "avg_coherence": avg_scores.get("coherence", 0.0),
            "avg_factuality": avg_scores.get("factuality", 0.0),
            "avg_overall": avg_scores.get("overall", 0.0),
            "total_flags": sum(flag_counts.values()),
            "flags_by_pillar": flag_counts,
            "judge_call_rate": 1.0 if judge is not None else 0.0,
            "judge_budget_used": 0,
            "judge_budget_total": 0,
            "total_samples": total_samples,
            "analysis_time": time.time() - start_time,
            "avg_time_per_sample": (time.time() - start_time) / total_samples if total_samples > 0 else 0.0
        }
        
        # Build final response
        analysis_result = {
            "job_id": job_id,
            "per_sample": per_sample,
            "summary": summary,
            "analysis_method": "pillars_v2",
            "config": {"judge_available": judge is not None},
            "timestamp": time.strftime("%Y-%m-%d %H:%M:%S")
        }
        
        # Print final summary
        print(f"🎉 CoT analysis complete! Processed {len(results)} samples")
        successful_samples = len([r for r in results if 'error' not in r])
        print(f"✅ Successfully analyzed: {successful_samples}/{len(results)} samples")
        if valid_results:
            print(f"📊 Average scores: {avg_scores}")
            print(f"🚩 Total flags found: {sum(flag_counts.values())}")
        print()
        
        return analysis_result
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error performing CoT analysis: {str(e)}")

@app.post("/jobs/{job_id}/cot-analysis/compute")
async def compute_cot_analysis(job_id: str):
    """Trigger CoT analysis computation and cache results"""
    try:
        # This could be enhanced to run analysis in background and cache results
        # For now, just redirect to the GET endpoint
        return await get_cot_analysis(job_id)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error computing CoT analysis: {str(e)}")

@app.post("/config")
async def save_configuration(request: dict):
    """Save configuration settings"""
    try:
        # Update the path manager configuration
        if 'openai_api_key' in request:
            path_manager._config.openai_api_key = request['openai_api_key']
            path_manager._save_config()
            print(f"✅ OpenAI API key updated in configuration")
        
        return {"message": "Configuration saved successfully"}
        
    except Exception as e:
        print(f"❌ Error saving configuration: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to save configuration: {str(e)}")

@app.post("/config/test-openai-key", response_model=OpenAITestResponse)
async def test_openai_key(request: OpenAITestRequest):
    """Test OpenAI API key validity"""
    try:
        # Import OpenAI here to avoid import errors if not available
        try:
            from openai import OpenAI
        except ImportError:
            return OpenAITestResponse(
                valid=False,
                error="OpenAI library not installed. Please install with: pip install openai"
            )
        
        # Test the API key by making a simple request
        client = OpenAI(api_key=request.api_key)
        
        try:
            # Make a simple test request to list models
            response = client.models.list()
            
            # Get basic info about available models
            model_info = {
                "total_models": len(response.data),
                "gpt_models": [m.id for m in response.data if "gpt" in m.id.lower()],
                "available_for_chat": [m.id for m in response.data if "gpt-4" in m.id or "gpt-3.5" in m.id]
            }
            
            return OpenAITestResponse(
                valid=True,
                model_info=model_info
            )
            
        except Exception as api_error:
            return OpenAITestResponse(
                valid=False,
                error=f"API key test failed: {str(api_error)}"
            )
            
    except Exception as e:
        return OpenAITestResponse(
            valid=False,
            error=f"Error testing API key: {str(e)}"
        ) 
@app.post("/jobs/{job_id}/truncation-analysis", response_model=TruncationAnalysisResponse)
async def run_truncation_analysis(job_id: str, request: TruncationAnalysisRequest):
    """Run CoT truncation analysis for a completed job"""
    start_time = time.time()
    
    try:
        if job_id not in job_db:
            raise HTTPException(status_code=404, detail="Job not found")
        
        job_info = job_db[job_id]
        result_file = job_info.get("result_file")
        
        if not result_file:
            raise HTTPException(status_code=404, detail="No result file found for this job")
        
        result_path = Path(result_file)
        if not result_path.exists():
            raise HTTPException(status_code=404, detail="Result file not found on disk")
        
        # Security check
        config = path_manager.get_config()
        allowed_dirs = [Path(config.output_dir), Path(config.logs_dir), Path(config.evaluation_dir)]
        
        is_allowed = False
        for allowed_dir in allowed_dirs:
            try:
                result_path.relative_to(allowed_dir)
                is_allowed = True
                break
            except ValueError:
                continue
        
        if not is_allowed:
            raise HTTPException(status_code=403, detail="Access denied to this file path")
        
        # Load the JSONL results
        samples = []
        with open(result_path, 'r') as f:
            for line in f:
                line = line.strip()
                if line:
                    try:
                        samples.append(json.loads(line))
                    except json.JSONDecodeError:
                        continue
        
        if not samples:
            raise HTTPException(status_code=400, detail="No valid samples found in result file")
        
        # Build truncation analysis command
        python_bin = config.python_path
        eval_dir = Path(config.evaluation_dir)
        truncation_script = eval_dir / "run_truncation_analysis_with_logs.py"
        
        if not truncation_script.exists():
            raise HTTPException(status_code=500, detail=f"Truncation analysis script not found: {truncation_script}")
        
        # Create output directory for this analysis
        output_dir = result_path.parent / "truncation_analysis"
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Create temporary input file for the analysis
        temp_input_file = output_dir / f"temp_input_{job_id}.jsonl"
        with open(temp_input_file, 'w') as f:
            for sample in samples:
                f.write(json.dumps(sample) + '\n')
        
        # Handle different backends
        if request.backend == "local":
            # Run locally
            cmd = [
                python_bin,
                str(truncation_script),
                "--input_file", str(temp_input_file),
                "--model_name_or_path", request.model_name_or_path,
                "--dataset_name", request.dataset_name,
                "--output_dir", str(output_dir),
                "--temperature", str(request.temperature),
                "--top_p", str(request.top_p),
            ]
            
            try:
                result = subprocess.run(cmd, capture_output=True, text=True, cwd=eval_dir, timeout=3600)  # 1 hour timeout
                if result.returncode != 0:
                    raise HTTPException(status_code=500, detail=f"Truncation analysis failed: {result.stderr}")
                
                computation_time = time.time() - start_time
                return _build_truncation_response(job_id, request, output_dir, computation_time)
                
            except subprocess.TimeoutExpired:
                raise HTTPException(status_code=500, detail="Truncation analysis timed out after 1 hour")
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"Error running truncation analysis: {str(e)}")
        
        elif request.backend == "slurm":
            # Submit to SLURM
            truncation_job_id = f"trunc_{job_id}_{int(time.time())}"
            
            # Create SLURM script
            scripts_dir = Path(config.scripts_dir)
            scripts_dir.mkdir(parents=True, exist_ok=True)
            
            script_path = scripts_dir / f"run_truncation_{truncation_job_id}.sh"
            sbatch_path = scripts_dir / f"job_truncation_{truncation_job_id}.sbatch"
            
            # Build the command for the script
            cmd = [
                python_bin,
                str(truncation_script),
                "--input_file", str(temp_input_file),
                "--model_name_or_path", request.model_name_or_path,
                "--dataset_name", request.dataset_name,
                "--output_dir", str(output_dir),
                "--temperature", str(request.temperature),
                "--top_p", str(request.top_p),
                "--job_id", truncation_job_id,  # Add job ID for unique filenames
            ]
            
            # Create the shell script
            escaped_cli = []
            for arg in cmd:
                escaped_cli.append(shlex.quote(str(arg)))
            
            script_content = f"""#!/bin/bash
cd {eval_dir}

# Set Hugging Face cache to work directory

# Fix MKL threading conflict
export MKL_THREADING_LAYER=GNU
export OMP_NUM_THREADS=1
export MKL_NUM_THREADS=1

# Activate conda environment
source {config.conda_env_path}/etc/profile.d/conda.sh
conda activate mathevalUI

# Run the truncation analysis
{' '.join(escaped_cli)}
"""
            
            script_path.write_text(script_content)
            script_path.chmod(0o755)
            
            # Create SLURM sbatch script with GPU allocation
            out_file_pattern = os.path.join(config.logs_dir, f"truncation-{truncation_job_id}-%j.out")
            err_file_pattern = os.path.join(config.logs_dir, f"truncation-{truncation_job_id}-%j.err")
            
            sbatch_content = f"""#!/bin/bash
#SBATCH -J truncation-{truncation_job_id}   # Job name
#SBATCH -o {out_file_pattern}      # Name of stdout output file (uses %j)
#SBATCH -e {err_file_pattern}      # Name of stderr error file (uses %j)
#SBATCH -p {config.slurm_partition}              # Queue (partition) name
#SBATCH -N 1                    # Total # of nodes
#SBATCH -n 1                    # Total # of tasks (single process for all GPUs)
#SBATCH -t {config.slurm_wall_time}              # Run time (hh:mm:ss)
#SBATCH --mail-type=all         # Send email at begin and end of job
#SBATCH -A {config.slurm_account}             # Project/Allocation name

# Fix MKL threading conflict
export MKL_THREADING_LAYER=GNU
export OMP_NUM_THREADS=1
export MKL_NUM_THREADS=1

export CUDA_VISIBLE_DEVICES=${{CUDA_VISIBLE_DEVICES:-0}}

{script_path}
"""
            
            sbatch_path.write_text(sbatch_content)
            
            try:
                result = subprocess.run(
                    ["sbatch", str(sbatch_path)],
                    capture_output=True,
                    text=True,
                    cwd=scripts_dir
                )
                
                if result.returncode == 0:
                    slurm_jid = result.stdout.strip().split()[-1]
                    
                    # Create actual file paths by replacing %j with SLURM job ID
                    actual_out_file = out_file_pattern.replace("%j", str(slurm_jid))
                    actual_err_file = err_file_pattern.replace("%j", str(slurm_jid))
                    
                    # Store the truncation job info
                    truncation_job_id_full = f"truncation_{job_id}_{slurm_jid}"
                    job_db[truncation_job_id_full] = {
                        "status": "queued",
                        "request": request.dict(),
                        "parent_job_id": job_id,
                        "run_path": str(script_path),
                        "sbatch_path": str(sbatch_path),
                        "backend": "slurm",
                        "slurm_jid": slurm_jid,
                        "out_file": actual_out_file,
                        "err_file": actual_err_file,
                        "output_dir": str(output_dir),
                        "temp_input_file": str(temp_input_file),
                        "dataset_name": request.dataset_name,
                        "model_name_or_path": request.model_name_or_path,
                        "temperature": request.temperature,
                        "top_p": request.top_p,
                        "start_time": start_time
                    }
                    save_job_db()
                    
                    return TruncationAnalysisResponse(
                        job_id=job_id,
                        status="queued",
                        message=f"Truncation analysis queued on SLURM with job ID {slurm_jid}",
                        raw_curves_path=None,
                        correct_plot_path=None,
                        incorrect_plot_path=None,
                        computation_time=None
                    )
                else:
                    raise HTTPException(status_code=500, detail=f"Failed to submit SLURM job: {result.stderr}")
                    
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"Error submitting to SLURM: {str(e)}")
        
        else:
            raise HTTPException(status_code=400, detail=f"Unsupported backend: {request.backend}")
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error performing truncation analysis: {str(e)}")

@app.get("/jobs/{job_id}/truncation-analysis/plot")
async def get_truncation_plot(job_id: str, plot_type: str = "correct"):
    """Get truncation analysis plot files"""
    if plot_type not in ["correct", "incorrect"]:
        raise HTTPException(status_code=400, detail="plot_type must be 'correct' or 'incorrect'")
    
    if job_id not in job_db:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job_info = job_db[job_id]
    result_file = job_info.get("result_file")
    
    if not result_file:
        raise HTTPException(status_code=404, detail="No result file found for this job")
    
    result_path = Path(result_file)
    output_dir = result_path.parent / "truncation_analysis"
    
    # Look for plot files
    plot_files = list(output_dir.glob(f"*_{plot_type}_*.png"))
    if not plot_files:
        raise HTTPException(status_code=404, detail=f"No {plot_type} plot found for this job")
    
    # Return the most recent plot file
    latest_plot = max(plot_files, key=lambda p: p.stat().st_mtime)
    
    # Security check
    config = path_manager.get_config()
    allowed_dirs = [Path(config.output_dir), Path(config.logs_dir), Path(config.evaluation_dir)]
    
    is_allowed = False
    for allowed_dir in allowed_dirs:
        try:
            latest_plot.relative_to(allowed_dir)
            is_allowed = True
            break
        except ValueError:
            continue
    
    if not is_allowed:
        raise HTTPException(status_code=403, detail="Access denied to this file path")
    
    return FileResponse(path=str(latest_plot), filename=latest_plot.name, media_type='image/png') 
